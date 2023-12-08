import logging
import re
import string
import time
from dataclasses import dataclass
from datetime import datetime
from itertools import chain, zip_longest
from typing import Any, Optional
import dateutil.parser
import pandas as pd
import numpy as np
import gspread
import gspread.utils
import gspread_dataframe
import gspread_formatting as gsformat
from gspread_formatting import CellFormat, TextFormat, format_cell_ranges, Link, Color as ColorFormatting
from google_workers.api import auth, get_service
from googleapiclient.discovery import build
from gspread.exceptions import CellNotFound, APIError
from shapely import wkt
from enum import Enum, auto
from colour import Color
from threading import get_ident, Lock
from openpyxl.utils.cell import get_column_letter
from pathlib import Path

class GoogleSheetRowSearchStrategy(Enum):
    CACHE = auto()
    REQUEST = auto()


class GoogleSheetWorker:
    BATCH_UPLOAD_SIZE = 5000
    SPREADSHEETS_CACHE_FOLDER = (Path(__file__).parent / '_spreadsheet_cache').resolve()
    SPREADSHEETS_CACHE_FOLDER.mkdir(exist_ok=True)

    def __init__(
            self,
            spread_url=None,
            spread_id=None,
            sheet_id=None,
            search_strategy=GoogleSheetRowSearchStrategy.CACHE,
            aliases=None,
            header_row=1,
            credentials=None
            ):
        self.logger = logging.getLogger(self.__class__.__name__)

        self._dataframe = None
        self._lock = Lock()
        self._raw_grid = None
        self._last_refresh_dataframe_time = 0
        self.datetime_format = '%Y-%m-%d %H:%M:%S'
        self._search_strategy = None
        self.search_strategy = search_strategy
        self.aliases = aliases
        self.header_row = header_row

        self.credentials = credentials if credentials is not None else auth()
        self.gspread_client = gspread.authorize(self.credentials)
        self.api_service = get_service('sheets')

        if spread_url:
            self.spread = self.gspread_client.open_by_url(spread_url)
            sheet_id = re.findall(r'gid=(\d+)', spread_url)[0]
        elif spread_id:
            self.spread = self.gspread_client.open_by_key(spread_id)
        else:
            raise TypeError('Expected one of parameter [spread_id, spread_url]')
        self.sheet_name = None
        if sheet_id:
            for w in self.spread.worksheets():
                if w.id == int(sheet_id):
                    self.sheet_name = w.title
                    break
            else:
                raise TypeError(f'Not found sheet with id={sheet_id}')
        self.refresh_sheet()


    def __repr__(self):
        return f'<GoogleSheetWorker(spread_id={self.spread_id}, sheet_name={self.sheet_name})>'

    def __define_strategy(self, strat):
        if strat == GoogleSheetRowSearchStrategy.CACHE:
            self.find_rows_by_values = self._find_rows_by_cache
            self.get_unique_values = self._get_unique_by_cache
        elif strat == GoogleSheetRowSearchStrategy.REQUEST:
            self.find_rows_by_values = self._find_rows_by_request
            self.get_unique_values = self._get_unique_by_request

    @property
    def search_strategy(self):
        return self._search_strategy

    @search_strategy.setter
    def search_strategy(self, strat):
        self.__define_strategy(strat)
        self._search_strategy = strat

    @property
    def reverse_aliases(self):
        if self.aliases:
            return {v: k for k, v in self.aliases.items()}

    @property
    def spread_id(self):
        return self.spread.id

    @property
    def raw_grid(self):
        if self._raw_grid is None:
            self.update_raw_grid()
        return self._raw_grid

    @property
    def color_dataframe(self):
        df = self.raw_grid.get_dataframe_by_property('background_color')
        df = df.drop(0, axis=0)
        df.index += 1
        rows, cols = self.dataframe.shape
        df = df.iloc[:rows, :cols]
        df = df.applymap(lambda x: Color('#'+x), na_action='ignore')
        df.columns = self.get_headers()[:len(df.columns)]
        return df

    @property
    def url(self):
        return f'https://docs.google.com/spreadsheets/d/{self.spread_id}/edit#gid={self.sheet_id}'

    @staticmethod
    def generate_hyperlink_format(hyperlink):
        cell_format = CellFormat(
            textFormat=TextFormat(
                link=Link(hyperlink),
                underline=True,
                foregroundColor=ColorFormatting.fromHex('#1155cc')
            ),
            hyperlinkDisplayType='LINKED',
        )
        return cell_format

    def _get_cache_filepath(self) -> Path:
        return self.SPREADSHEETS_CACHE_FOLDER / f'{self.spread_id}_{self.sheet_id}.pickle'

    def _save_cache(self) -> None:
        self.logger.debug('Save cache...')
        cache_path = self._get_cache_filepath()
        cache_path.parent.mkdir(exist_ok=True)
        self._dataframe.to_pickle(cache_path)

    def _load_cache(self) -> Optional[pd.DataFrame]:
        self.logger.debug('Read cache...')
        cache_path = self._get_cache_filepath()
        if cache_path.exists():
            df = pd.read_pickle(cache_path)
            return df

    def refresh_sheet(self):
        try:
            if self.sheet_name:
                self._sheet = self.spread.worksheet(self.sheet_name)
            else:
                self._sheet = self.spread.sheet1
            return self._sheet
        except APIError as e:
            if e.response.json()['error'].get('code') == 429:
                self.logger.debug('Quota exceeded')
                time.sleep(10)
                return self.refresh_sheet()
            else:
                raise e

    @property
    def sheet(self):
        return self._sheet

    @property
    def sheet_id(self):
        return self.sheet.id

    @property
    def first_row(self):
        return self.header_row + 1

    def replace_headers(self, values):
        self.sheet.delete_row(self.header_row)
        self.sheet.insert_row(values, self.header_row)

    def set_headers(self, values):
        self.sheet.insert_row(values, self.header_row)

    def get_headers(self):
        r = self.sheet.get(f'{self.header_row}:{self.header_row}')
        return r[0]

    def get_aliased_headers(self):
        a = self.reverse_aliases
        return [a[k] if k in a else k for k in self.get_headers()]

    def get_worksheet_filters(self):
        r = self.api_service.get(spreadsheetId=self.spread_id).execute()
        ws = [i for i in r['sheets'] if i['properties']['title'] == self.sheet_name]
        if ws:
            return ws[0].get('filterViews', [])
        else:
            raise TypeError(f'Не найден лист с именем "{self.sheet_name}"')

    def delete_sheet_filters(self, filter_ids):
        requests = [{'deleteFilterView': {'filterId': i}} for i in filter_ids]
        body = {'requests': requests}
        r = self.api_service.batchUpdate(spreadsheetId=self.spread_id, body=body).execute()
        return r

    def delete_all_sheet_filters(self):
        self.logger.debug('Deleting existing filters')
        r = self.api_service.get(spreadsheetId=self.spread_id).execute()
        ids = [i.get('filterViews', []) for i in r['sheets']]
        ids = list(chain(*ids))
        ids = [i.get('filterViewId', None) for i in ids]
        ids = [i for i in ids if i]
        if ids:
            r = self.delete_sheet_filters(ids)
            return r

    def batch_update(self, requests):
        self.logger.debug('Batch update')
        body = {'requests': requests}
        response = self.api_service.batchUpdate(spreadsheetId=self.spread_id, body=body).execute()
        return response

    def _need_to_update_dataframe(self):
        conditions = []
        conditions.append(self.sheet.row_count != self._dataframe.shape[0] + self.header_row)
        # conditions.append(self.sheet.col_count != self._dataframe.shape[1])
        return any(conditions)

    def update_dataframe(self):
        self.logger.debug(f'Update dataframe by thread {get_ident()}')
        self.refresh_sheet()
        # собираем всю таблицу если нет определенных колонок
        if self.aliases is None:
            # gspread_dataframe уродует типы данных, поэтому делаем все сами
            values = self.sheet.get_all_values()
        # в противном случае кэшируем только нужные колонки для экономии памяти
        else:
            # находим индексы нужных нам колонок
            headers = self.get_headers()
            req_ranges = []
            for alias_col in self.aliases.values():
                icol = headers.index(alias_col) + 1
                acol = get_column_letter(icol)
                req_ranges.append(f'{acol}:{acol}')
            # собираем данные по колонкам одним запросом
            raw_values = self.sheet.batch_get(req_ranges)
            # заполняем пустые массивы, чтобы они не пропадали в chain
            # не сбивая тем самым матрицу
            for cv in raw_values:
                for rv in cv:
                    if len(rv) == 0:
                        rv.append('')
            # форматируем данные
            values = []
            for row in zip_longest(*raw_values, fillvalue=['']):
                values.append(list(chain.from_iterable(row)))
        df = pd.DataFrame(values)
        df.index += 1
        df.columns = df.iloc[self.header_row-1]
        df.columns.name = None
        df = df.iloc[self.header_row:]
        df = df.replace('', np.nan)
        self._dataframe = df
        self._last_refresh_dataframe_time = time.time()
        self._save_cache()
        self.logger.debug(f'Finish to update dataframe by thread {get_ident()}')

    @property
    def dataframe(self):
        with self._lock:
            if self._dataframe is None:
                cache_df = self._load_cache()
                if cache_df is None:
                    self.update_dataframe()
                else:
                    self._dataframe = cache_df
            # todo возможно нужно убрать, так как логика с сопоставлением размера таблицы корявая, перепроверка кэша справляется лучше
            # if self._need_to_update_dataframe():
            #     self.update_dataframe()
            return self._dataframe

    @property
    def aliased_dataframe(self):
        if self.reverse_aliases:
            df = self.dataframe.rename(self.reverse_aliases, axis=1)
            return df.loc[:, self.aliases]
        else:
            self.logger.debug('Alises not setted. Return original dataframe')
            return self.dataframe

    def upload_dataframe(self, gdf, start_row_index=0):
        self.refresh_sheet()  # необходимо актуализировать перед изменениями
        # подготавливаем данные - строки больше 5000 недопустимы
        for ic, c in enumerate(gdf.columns):
            series = gdf[c].dropna()
            if len(series) > 0:
                # с этого момента я проклинаю языки с динамической типизацией (а они меня)
                sdt = str(series.dtype)
                if 'datetime' in sdt:
                    gdf[c] = gdf[c].dt.strftime(self.datetime_format)
                elif 'object' in sdt:
                    try:
                        gdf[c] = gdf[c].str[:5000]
                    except AttributeError:
                        pass
        for i in range(0, len(gdf), self.BATCH_UPLOAD_SIZE):
            self.logger.debug(f'Upload dataframe: {i} - {i + self.BATCH_UPLOAD_SIZE} [{len(gdf)}]')
            current_start_row = self.first_row + start_row_index + i
            if self.sheet.row_count < current_start_row:
                self.sheet.add_rows(current_start_row - self.sheet.row_count)
            gspread_dataframe.set_with_dataframe(self.sheet, gdf.iloc[i:i + self.BATCH_UPLOAD_SIZE], row=current_start_row,
                                                 include_column_header=False, include_index=False, resize=False)

    def format_table(self, boolean_сolumns=[]):
        self.logger.info('Format table')
        ws = self.sheet
        header_format = gsformat.cellFormat(
            backgroundColor=gsformat.color(0.9, 0.9, 0.9),
            textFormat=gsformat.textFormat(bold=True),
            horizontalAlignment='CENTER'
        )
        cell_format = gsformat.cellFormat(
            backgroundColor=gsformat.color(1.0, 1.0, 1.0),
            textFormat=gsformat.textFormat(bold=False),
            horizontalAlignment='LEFT'
        )
        gsformat.format_cell_range(self.sheet, f"{self.header_row}:{self.sheet.row_count}", cell_format)
        gsformat.format_cell_range(self.sheet, f"{self.header_row}:{self.header_row}", header_format)
        for boolean_сolumn in boolean_сolumns:
            col_index = self.get_aliased_headers().index(boolean_сolumn) + 1
            a1_cell = gspread.utils.rowcol_to_a1(1, col_index)
            col_label = re.findall(r'[A-Z]+', a1_cell)[0]
            boolean_range = f'{col_label}:{col_label}'
            validation_rule = gsformat.DataValidationRule(
                gsformat.BooleanCondition('BOOLEAN', []),
                showCustomUi=True
            )
            gsformat.set_data_validation_for_cell_range(self.sheet, boolean_range, validation_rule)

    ### start: SQL methods

    def clear_worksheet(self):
        '''SQL drop implement'''
        ws = self.sheet
        ws.delete_rows(start_index=1, end_index=ws.row_count - 1)
        ws.delete_columns(start_index=1, end_index=ws.col_count - 1)
        ws.update_cell(1, 1, '')
        gsformat.set_data_validation_for_cell_range(ws, "A1", None)

    def truncate_worksheet(self):
        '''SQL truncate implement'''
        # удаляем все строки кроме заголовка
        self.logger.info('Truncate table')
        ws = self.sheet
        if ws.row_count > 1:
            ws.delete_rows(start_index=2, end_index=ws.row_count)

    def value_formatter(self, v):
        if isinstance(v, datetime):
            return v.strftime(self.datetime_format)
        elif isinstance(v, str):
            if re.match(r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\+\d{2}:\d{2}', v):
                v = dateutil.parser.parse(v)
                return v.strftime(self.datetime_format)
            else:
                return v
        elif v is None:
            return ''
        else:
            return v

    def find_column(self, col_name: str):
        self.logger.debug(f'Search column {col_name}')
        if self.aliases:
            col_name = self.aliases[col_name]
        return self.get_headers().index(col_name) + 1

    def find_rows_by_value(self, col_value: str, row_value: str):
        '''
        Get row indexes which contain value in column
        '''
        col_index = self.find_column(col_value)
        cells = self.sheet.findall(str(row_value), in_column=col_index)
        return [i.row for i in cells]

    def _find_rows_by_request(self, row_values: dict):
        '''Get row indexes which contain values in many column in same time'''
        self.logger.debug(f'Finding rows by values {row_values} [request]')
        row_sets = []
        for k, v in row_values.items():
            rows = self.find_rows_by_value(k, v)
            row_sets.append(set(rows))
        return set.intersection(*row_sets)

    def _find_rows_by_cache(self, row_values: dict, **kwargs):
        self.logger.debug(f'Finding rows by values {row_values} [cache]')
        after_cache_update = kwargs.get('after_cache_update')
        rows = set(self.aliased_dataframe.loc[(
                           self.aliased_dataframe[row_values.keys()].astype(str) == pd.Series(row_values).astype(str)
                       ).all(axis=1)
                   ].index)
        #check cache is valid
        for row in rows:
            for k,v in row_values.items():
                cell_value = self.sheet.cell(row, self.find_column(k)).value
                if str(cell_value) != str(v):
                    self.logger.info(f'Cell values mismatch {v} - {cell_value}. Updating dataframe and repeat...')
                    self.update_dataframe()
                    return self._find_rows_by_cache(row_values, after_cache_update=True)
        if len(rows) == 0:
            if after_cache_update:
                raise CellNotFound(f'Not found row with values {row_values}')
            else:
                self.logger.debug('Cell not found, update cache')
                self.update_dataframe()
                return self._find_rows_by_cache(row_values=row_values, after_cache_update=True)
        return rows

    def update_row_by_id(self, id_values: dict, update_values: dict, **kwargs):
        self.logger.debug(f'Updating rows {id_values} - {update_values}')
        id_rows = self.find_rows_by_values(id_values)
        for k,v in update_values.items():
            value_col = self.find_column(k)
            for value_row in id_rows:
                v = self.value_formatter(v)
                self.sheet.update_cell(value_row, value_col, v)

    def insert_row_by_id(self, insert_values):
        # todo обновить функцию
        raise NotImplementedError()

    def delete_row_by_id(self, id_values):
        # todo проверить корректность удаления (смещение массива)
        raise NotImplementedError()

    def _get_unique_by_request(self, column_index, include_header=False):
        v = self.sheet.col_values(column_index)
        if not include_header:
            v = v[1:]
        return set(v)

    def _get_unique_by_cache(self, column_index, include_header=False):
        v = self.aliased_dataframe.iloc[:, column_index+1].unique()
        if include_header:
            col = self.aliased_dataframe.columns[column_index+1]
            v = list(v).insert(0, col)
        return set(v)

    ### end: SQL methods

    def generate_filter_request(self, filter_name, sql):
        headers = self.get_aliased_headers()
        equal_operators = {
            '=': [
                (lambda x: isinstance(x, (int, float)), 'NUMBER_EQ'),
                (lambda x: '::date' in x, 'DATE_EQ'),
                (lambda x: isinstance(x, str), 'TEXT_EQ'),
            ],
            '!=': [
                (lambda x: isinstance(x, (int, float)), 'NUMBER_NOT_EQ'),
                (lambda x: '::date' in x, 'DATE_NOT_EQ'),
                (lambda x: isinstance(x, str), 'TEXT_NOT_EQ'),
            ],
        }
        number_operators = {
            '<': 'NUMBER_LESS',
            '>': 'NUMBER_GREATER',
            '>=': 'NUMBER_GREATER_THAN_EQ',
            '<=': 'NUMBER_LESS_THAN_EQ',
        }
        date_operators = {
            '<': 'DATE_BEFORE',
            '>': 'DATE_AFTER',
            '>=': 'DATE_ON_OR_AFTER',
            '<=': 'DATE_ON_OR_BEFORE',
        }
        logic_operators = ['LIKE']
        array_operators = ['NOT IN', 'IN']

        sql = re.sub(r'\s{2,}', ' ', sql)
        sql = sql.split('WHERE')[-1].strip()
        if 'ORDER BY' in sql:
            raise NotImplementedError('order by')
        if 'OR' in sql:
            raise TypeError('OR statement is not supported')
        sql_filters = sql.split(' AND ')
        filter_specs = []
        for sql_filter_string in sql_filters:
            # определяем колонку и ее индекс
            criteria_column = sql_filter_string.split(' ')[0]
            sql_filter_string = sql_filter_string.replace(criteria_column, '').strip()
            criteria_column_index = headers.index(criteria_column)

            # определяем оператор
            operator_pat = re.compile('|'.join(
                logic_operators + array_operators + list(date_operators.keys()) + list(equal_operators.keys())))
            operator = operator_pat.findall(sql_filter_string)
            sql_filter_string = operator_pat.sub('', sql_filter_string).strip("' ()'")
            if operator:
                operator = operator[0]
                if operator in array_operators:
                    if operator in ('NOT IN', 'IN'):
                        criteria_operator = operator
                        criteria_method = 'hiddenValues'
                elif operator in logic_operators:
                    criteria_operator = 'TEXT_CONTAINS'
                    criteria_method = 'condition'
                elif operator in equal_operators:
                    criteria_method = 'condition'
                    for option in equal_operators[operator]:
                        if option[0](sql_filter_string):
                            criteria_operator = option[1]
                            break
                else:
                    criteria_method = 'condition'
                    if '::date' in sql_filter_string:
                        criteria_operator = date_operators[operator]
                    else:
                        criteria_operator = number_operators[operator]
            else:
                criteria_operator = 'SHOW_TRUE'
                criteria_method = 'hiddenValues'

            if criteria_method == 'hiddenValues':
                list_values = re.split(r"'[,\s]+'", sql_filter_string)
                if criteria_operator == 'SHOW_TRUE':
                    criteria_value = ['ЛОЖЬ']
                elif criteria_operator == 'NOT IN':
                    criteria_value = list_values
                elif criteria_operator == 'IN':
                    unique_values = self.get_unique_values(criteria_column_index + 1, include_header=False)
                    diff = list(unique_values.difference(set(list_values)))
                    criteria_value = diff
                else:
                    raise Exception(f'criteria_method == hiddenValues: lost case - {criteria_operator}')

            elif criteria_method == 'condition':
                if criteria_operator == 'TEXT_CONTAINS':
                    criteria_value = {'type': criteria_operator,
                                      'values': [{'userEnteredValue': sql_filter_string.strip('% ')}]}
                elif criteria_operator == 'ONE_OF_LIST':
                    criteria_value = {'type': criteria_operator,
                                      'values': [{'userEnteredValue': i.strip("' ")} for i in
                                                 sql_filter_string.split(',')]
                                      }
                else:
                    sql_filter_string = re.sub("'::date", '', sql_filter_string)
                    criteria_value = {'type': criteria_operator, 'values': [{'userEnteredValue': sql_filter_string}]}

            filter_spec = {'columnIndex': criteria_column_index, 'filterCriteria': {criteria_method: criteria_value}}
            filter_specs.append(filter_spec)
        self.logger.debug(f'Filter: {str(filter_specs)}')
        request = {
            'addFilterView': {
                'filter': {
                    'title': filter_name,
                    'range': {'sheetId': 0},
                    'filterSpecs': filter_specs,
                }
            }
        }
        return request

    def extract_column_hyperlink(self, icol):
        acol = get_column_letter(icol)
        return self.extract_hyperlink(f'{acol}:{acol}')

    def extract_hyperlink(self, cell_range):
        r = self.api_service.get(spreadsheetId=self.spread_id, ranges=f"'{self.sheet_name}'!{cell_range}",
                                 includeGridData=True).execute()
        hyperlinks = []
        for row_data in r['sheets'][0]['data'][0]['rowData']:
            if 'values' in row_data:
                rows_hyperlinks = [i.get('hyperlink', None) for i in row_data['values']]
            else:
                rows_hyperlinks = [None]
            hyperlinks.append(rows_hyperlinks)
        return hyperlinks

    def get_hyperlink_dataframe(self):
        self.logger.debug(f'Download spreadsheet hyper dataframe...')
        resp_step = 1000
        total_rows, total_cols = self.dataframe.shape
        first_acol = 'A'
        last_acol = get_column_letter(total_cols)
        first_row = self.first_row
        last_row = self.header_row + total_rows
        hypers = []
        for i in range(first_row-1, last_row, resp_step):
            r = f'{first_acol}{i + 1}:{last_acol}{i + resp_step}'
            self.logger.debug(f'Hyperlinks parse: {r}')
            hypers += self.extract_hyperlink(r)
        hdf = pd.DataFrame(hypers, columns=self.get_headers())
        hdf.index += self.first_row
        return hdf
    
    def set_column_hyperlinks(self, icol, hyperlinks):
        acol = get_column_letter(icol)
        self.set_hyperlinks(f'{acol}:{acol}', hyperlinks)

    def set_hyperlinks(self, cell_range, hyperlinks):
        updates = []
        cells = self.sheet.range(cell_range)
        assert len(cells) == len(hyperlinks)
        updates = [(cells[i].address, self.generate_hyperlink_format(hyperlinks[i])) for i in range(len(cells))]
        format_cell_ranges(self.sheet, updates)

    def set_hyperlinks_dataframe(self, hdf):
        updates = []
        for c in hdf.columns:
            if len(hdf[c].dropna()) > 0:
                icol = self.find_column(c.replace('#href', ''))
                acol = get_column_letter(icol)
                for i,v in hdf[c].dropna().iteritems():
                    updates.append((f'{acol}{i}', self.generate_hyperlink_format(v)))
        format_cell_ranges(self.sheet, updates)

    def update_raw_grid(self):
        self.logger.debug('Init raw grid')
        cells = []
        for i in range(1, self.sheet.row_count, self.BATCH_UPLOAD_SIZE):
            r = self.api_service.get(
                spreadsheetId=self.spread_id,
                ranges=f"'{self.sheet_name}'!{i}:{i+self.BATCH_UPLOAD_SIZE}",
                includeGridData=True
            ).execute()
            data = r['sheets'][0]['data'][0]
            for row in data['rowData']:
                if row.get('values', None):
                    row_cells = [GSCell.from_json(cell) for cell in row['values']]
                    cells.append(row_cells)
        self._raw_grid = GSGrid(cells)
        return self._raw_grid


##########################################################################
# Здесь идет блок вспомогательных дата-классов, которые копирует         #
# Spreadsheets rest api ответы и упрощают взаимодействие с ответами rest # 
##########################################################################

@dataclass
class GSCell:
    user_entered_value: Any = None
    user_entered_type: str = None
    effective_value: Any = None
    effective_type: str = None
    formatted_value: str = None
    background_color: str = None
    hyperlink: str = None
    note: str = None

    @staticmethod
    def color_to_hex(red=1.0, green=1.0, blue=1.0) -> str:
        r = int(red * 255)
        g = int(green * 255)
        b = int(blue * 255)
        return '%02x%02x%02x' % (r, g, b)

    @classmethod
    def from_json(cls, data):
        uv = ut = ev = et = None
        for k, v in data.get('userEnteredValue', {}).items():
            ut, uv = k, v
        for k, v in data.get('effectiveValue', {}).items():
            et, ev = k, v
        return cls(
            user_entered_value=uv,
            user_entered_type=ut,
            effective_value=ev,
            effective_type=et,
            formatted_value=data.get('formattedValue', None),
            background_color=GSCell.color_to_hex(**data.get('effectiveFormat', {}).get('backgroundColor', {})),
            hyperlink=data.get('hyperlink', None),
            note=data.get('note', None),
        )


class GSGrid:
    def __init__(self, grid):
        self.grid = grid

    def get_dataframe_by_property(self, property_name):
        return pd.DataFrame([[getattr(cell, property_name) for cell in row] for row in self.grid])

    def __getattr__(self, item):
        if item in self.grid[0][0].__annotations__.keys():
            return self.get_dataframe_by_property(item)
        else:
            raise AttributeError(f'No such attribute {item}')


if __name__ == '__main__':
    gs = GoogleSheetWorker(spread_url='https://docs.google.com/spreadsheets/d/1-vcr_Ks6qAjWU9JaYN2rAnwH9f8sg2IHu1MPVJPi09o/edit#gid=0')
